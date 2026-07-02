"""
generate_master_timetable.py
─────────────────────────────────────────────────────────────────────────────
Reads output/timetable.json (produced by main.py) and builds a Master
Timetable spreadsheet where:

  - Columns = timeslots (08:00-08:50, 08:50-09:40, ... 18:00-18:50), header
    row merged only across the single header row (normal single row).
  - Rows = each Day gets as many rows as the MAX number of classes that run
    concurrently in any of its timeslots. The "Day" cell (column A) is
    merged vertically to span all of that day's rows.
  - Each individual class gets its OWN cell — cells are never merged
    together or stacked with newline-joined text. If slot X on Monday has
    3 simultaneous classes (different rooms) and slot Y only has 1, the
    Monday block still has rows for all 3, with slot Y's row 2 and row 3
    simply left blank.
  - The lunch break slot (13:00-13:50) is rendered as a single cell merged
    across that day's row-block, labelled "LUNCH BREAK".
  - 2-slot practicals show in full in their starting slot's row, and as a
    short "↳ contd." note in the same row of the following slot column
    (so it stays visually aligned with the class it belongs to).

Place this file in: backend_new/  (next to main.py, extractor.py, scheduler.py)

Usage:
    python generate_master_timetable.py
        → reads output/timetable.json, writes output/master_timetable.xlsx

    python generate_master_timetable.py path/to/timetable.json path/to/out.xlsx
        → custom input/output paths
"""

import json
import os
import sys
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TYPE_ABBREV = {"lecture": "LEC", "tutorial": "TUT", "practical": "PRAC"}


def load_timetable(json_path):
    with open(json_path, encoding="utf-8") as f:
        return json.load(f)


def format_class(c):
    typ = TYPE_ABBREV.get(c["type"], c["type"].upper())
    return (
        f"{c['course_code']} [{typ}]\n"
        f"{c['title']}\n"
        f"Room {c['room_id']} \u00b7 {c['teacher']}"
    )


def verify_counts(timetable, cell_classes, columns):
    """
    Cross-checks that every class in timetable['timetable'] was placed into
    exactly one (day, slot) cell — catches silent drops from bugs like a
    slot_index that doesn't map to any column, or a day name mismatch.
    Returns True if everything matches, prints details either way.
    """
    days = timetable["meta"]["days"]
    expected_total = len(timetable["timetable"])

    # Count what we actually queued into cell_classes.
    placed_total = sum(len(v) for v in cell_classes.values())

    print("\n  \u2014 Verification \u2014")
    ok = True

    # Catch classes whose "day" field doesn't match meta['days'] at all —
    # these get queued into cell_classes but will never be rendered because
    # the render loop only iterates over meta['days'].
    valid_days = set(days)
    bad_day_classes = [c for c in timetable["timetable"] if c["day"] not in valid_days]
    if bad_day_classes:
        ok = False
        bad_days_seen = sorted({c["day"] for c in bad_day_classes})
        print(f"  \u2717 {len(bad_day_classes)} class(es) have a 'day' value not in "
              f"meta['days'] {days} \u2014 they will NOT appear anywhere in the "
              f"sheet: {bad_days_seen}")

    if placed_total != expected_total:
        ok = False
        print(f"  \u2717 MISMATCH: {expected_total} classes in JSON but "
              f"{placed_total} were placed into cells.")
    else:
        print(f"  \u2713 Total classes: {placed_total}/{expected_total} placed.")

    # Per-day breakdown: JSON's own by_day (if present) vs what we queued.
    by_day_json = timetable.get("by_day")
    for day in days:
        json_count = (
            len(by_day_json[day]) if by_day_json and day in by_day_json
            else sum(1 for c in timetable["timetable"] if c["day"] == day)
        )
        placed_count = sum(
            len(cell_classes.get((day, i), []))
            for i in range(len(columns))
        )
        status = "\u2713" if json_count == placed_count else "\u2717"
        if json_count != placed_count:
            ok = False
        print(f"  {status} {day:10s}: {placed_count}/{json_count} classes placed")

    if not ok:
        print("  \u26a0  Some classes were not placed \u2014 check for slot_index "
              "values outside the expected day/slot range, or day names in "
              "the JSON that don't match meta['days'].")
    print()
    return ok


def build_master(timetable, output_path):
    meta          = timetable["meta"]
    days          = meta["days"]
    slots_per_day = meta["slots_per_day"]

    # meta["timeslots"] is a flat weekly list (day * slots_per_day + local).
    # The first `slots_per_day` entries give the day-template (start/end/
    # is_lunch are identical across all days).
    columns = meta["timeslots"][:slots_per_day]
    lunch_local_idx = next((i for i, s in enumerate(columns) if s.get("is_lunch")), None)

    # (day, local_slot_index) -> [class dicts starting there], in a stable
    # order (by unit_id) so re-runs produce the same layout.
    cell_classes = defaultdict(list)
    # (day, local_slot_index) -> [course_code of a 2-slot practical whose
    # SECOND half lands here]
    cell_continuation = defaultdict(list)

    for c in sorted(timetable["timetable"], key=lambda x: x["unit_id"]):
        day = c["day"]
        local_idx = c["slot_index"] % slots_per_day
        cell_classes[(day, local_idx)].append(c)
        if c["slots"] == 2:
            cell_continuation[(day, local_idx + 1)].append(c)

    verify_counts(timetable, cell_classes, columns=columns)

    # ── Determine how many rows each day's block needs ──
    # A row is needed per "concurrent slot" — the max, across all non-lunch
    # timeslots in that day, of (classes starting there) OR (a continuation
    # landing there, sharing a row with whichever class is already using
    # that row index if possible). We keep it simple: row count for a day
    # = max over local slots of len(cell_classes[(day, local_idx)]),
    # with continuations placed into existing rows (never adding a new row
    # just for a continuation, since a continuation always accompanies a
    # unit that already occupies a row from its starting slot).
    day_row_count = {}
    for day in days:
        max_rows = 1
        for local_idx, slot in enumerate(columns):
            if slot.get("is_lunch"):
                continue
            n = len(cell_classes.get((day, local_idx), []))
            if n > max_rows:
                max_rows = n
        day_row_count[day] = max_rows

    wb = Workbook()
    ws = wb.active
    ws.title = "Master Timetable"

    FONT_NAME = "Arial"
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(name=FONT_NAME, color="FFFFFF", bold=True, size=11)
    lunch_fill  = PatternFill("solid", fgColor="FFE0B2")
    lunch_font  = Font(name=FONT_NAME, italic=True, color="8A5A00", size=11)
    day_fill    = PatternFill("solid", fgColor="EDEDED")
    day_font    = Font(name=FONT_NAME, bold=True, size=12)
    body_font   = Font(name=FONT_NAME, size=9)
    contd_font  = Font(name=FONT_NAME, size=8, italic=True, color="777777")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_left   = Alignment(wrap_text=True, vertical="top", horizontal="left")
    center      = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # ── Header row (row 1) ──
    HEADER_ROW = 1
    corner = ws.cell(row=HEADER_ROW, column=1, value="Day \\ Time")
    corner.font, corner.fill, corner.alignment, corner.border = (
        header_font, header_fill, center, border,
    )
    for ci, slot in enumerate(columns, start=2):
        label = f"{slot['start']}\u2013{slot['end']}"
        if slot.get("is_lunch"):
            label += "\n(LUNCH)"
        cell = ws.cell(row=HEADER_ROW, column=ci, value=label)
        cell.font, cell.fill, cell.alignment, cell.border = (
            header_font, header_fill, center, border,
        )

    # ── Data rows ──
    current_row = HEADER_ROW + 1
    for day in days:
        n_rows = day_row_count[day]
        start_row = current_row
        end_row   = current_row + n_rows - 1

        # Day cell — merged across this day's row block.
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        day_cell = ws.cell(row=start_row, column=1, value=day)
        day_cell.font, day_cell.fill, day_cell.alignment = (day_font, day_fill, center)
        for r in range(start_row, end_row + 1):
            ws.cell(row=r, column=1).border = border

        for ci, slot in enumerate(columns, start=2):
            local_idx = ci - 2

            if slot.get("is_lunch"):
                # Merge the lunch cell across this day's row block too.
                ws.merge_cells(start_row=start_row, start_column=ci, end_row=end_row, end_column=ci)
                cell = ws.cell(row=start_row, column=ci, value="LUNCH BREAK")
                cell.fill, cell.font, cell.alignment = (lunch_fill, lunch_font, center)
                for r in range(start_row, end_row + 1):
                    ws.cell(row=r, column=ci).border = border
                continue

            classes = cell_classes.get((day, local_idx), [])
            conts   = cell_continuation.get((day, local_idx), [])

            for i in range(n_rows):
                r = start_row + i
                cell = ws.cell(row=r, column=ci)
                cell.border = border
                cell.alignment = wrap_left

                if i < len(classes):
                    cell.value = format_class(classes[i])
                    cell.font = body_font
                elif i < len(classes) + len(conts):
                    cont = conts[i - len(classes)]
                    cell.value = f"\u21b3 {cont['course_code']} (contd.)"
                    cell.font = contd_font
                # else: leave blank

        current_row = end_row + 1

    # ── Column widths / row heights ──
    ws.column_dimensions["A"].width = 14
    for ci in range(2, len(columns) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 26
    ws.row_dimensions[HEADER_ROW].height = 34
    for r in range(HEADER_ROW + 1, current_row):
        ws.row_dimensions[r].height = 85

    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=2).coordinate
    ws.sheet_view.showGridLines = False

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    total_rows = current_row - HEADER_ROW - 1
    print(f"\u2713 Master timetable saved \u2192 {output_path}")
    print(f"  Days              : {len(days)}")
    print(f"  Timeslots/day     : {len(columns)}")
    print(f"  Total data rows   : {total_rows}  (rows/day: "
          f"{ {d: day_row_count[d] for d in days} })")
    print(f"  Total class sessions : {len(timetable['timetable'])}")


if __name__ == "__main__":
    default_input  = os.path.join("output", "timetable.json")
    default_output = os.path.join("output", "master_timetable.xlsx")

    json_path   = sys.argv[1] if len(sys.argv) > 1 else default_input
    output_path = sys.argv[2] if len(sys.argv) > 2 else default_output

    if not os.path.exists(json_path):
        print(f"\n  \u2717 ERROR: Cannot find '{json_path}'.")
        print("  Run main.py first to generate timetable.json, or pass a path:")
        print("  python generate_master_timetable.py path/to/timetable.json path/to/out.xlsx\n")
        raise SystemExit(1)

    timetable = load_timetable(json_path)
    build_master(timetable, output_path)