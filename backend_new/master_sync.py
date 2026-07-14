"""
master_sync.py
Lets an admin hand-edit the *actual* master_timetable.xlsx (the merged-cell
grid produced by genMaster.build_master) and re-upload it to become the new
published timetable.json — no separate "editable" export format, no
re-running the scheduler.

Classes can be freely added, removed, or retyped (LEC/TUT/PRAC) in the
sheet — the sheet is the new source of truth for what classes exist. What
IS still checked is structural: every cell must parse, reference a real
room, and land on a real timeslot.

Who's enrolled in what is decided by course_code, not by exact unit_id.
extracted_data.json's enrollment_groups tells us which *courses* each
student group (semester|major|minor) takes. Any class session in the
uploaded sheet for one of those course_codes belongs to that group,
regardless of how many sessions that course has now or what type they
are. That's what makes adding/removing/retyping sessions safe — it
doesn't touch who's enrolled in the course itself.

Grid layout (produced by genMaster.build_master)
----------------------------------------------------
Column A is the day (merged vertically across that day's row-block),
columns 2..N are timeslots (header text like "08:00-08:50"), each class
gets its own cell as multi-line text:

    TB2401 [LEC]
    Intro to Databases
    Room LT101 . Dr. Rao

A 2-slot practical also drops a short "-> TB2401 (contd.)" marker in the
following slot's column, on the same row, purely so the grid reads
correctly — it isn't a second class, and is skipped on import.

Entry point
------------
  import_master_timetable(input_xlsx, extracted_json, output_json)
      Parses the uploaded grid xlsx, rebuilds timetable.json (classes /
      by_day / by_student), and returns the rebuilt dict.
"""

from __future__ import annotations

import json
import os
import re
from collections import defaultdict
from typing import Any

from openpyxl import load_workbook

from scheduler import DAYS as DAYS_ORDER, SLOTS_PER_DAY, ALL_SLOTS, _SLOT_MINUTES

TYPE_ABBREV = {"lecture": "LEC", "tutorial": "TUT", "practical": "PRAC"}
TYPE_FROM_ABBREV = {v: k for k, v in TYPE_ABBREV.items()}
TYPE_LETTER = {"lecture": "L", "tutorial": "T", "practical": "P"}

# "TB2401 [LEC]"
_HEADER_LINE_RE = re.compile(r"^\s*(\S+)\s*\[(\w+)\]\s*$")
# "Room LT101 \u00b7 Dr. Rao"  (the separator genMaster.py writes is U+00B7)
_ROOM_LINE_RE = re.compile(r"^\s*Room\s+(\S+)\s*\u00b7\s*(.+?)\s*$")

# One day's worth of the timeslot template, straight from scheduler.py.
_DAY_TEMPLATE = ALL_SLOTS[:SLOTS_PER_DAY]
_LUNCH_LOCAL_IDX = next((i for i, s in enumerate(_DAY_TEMPLATE) if s["is_lunch"]), None)


def _parse_header_time_label(label) -> "str | None":
    """Extract the start time (HH:MM) from a header cell like
    '08:00-08:50' or '13:00-13:50\\n(LUNCH)'. Returns None if it doesn't
    look like a time label."""
    if not label:
        return None
    first_line = str(label).split("\n")[0]
    parts = re.split(r"[\u2013\u2012\u2010-]", first_line)
    start = parts[0].strip()
    if re.fullmatch(r"\d{1,2}:\d{2}", start):
        return start
    return None


def _read_master_grid(input_xlsx: str) -> list:
    """Parse genMaster.py's merged-cell workbook into a flat list of
    {day, local_idx, course_code, unit_type, room_id, teacher} dicts — one
    per real class cell (continuation markers and blanks are skipped)."""
    wb = load_workbook(input_xlsx, data_only=True)
    sheet_name = "Master Timetable" if "Master Timetable" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    max_col = ws.max_column
    max_row = ws.max_row
    if max_row < 2 or max_col < 2:
        raise ValueError("Uploaded file doesn't look like a master timetable grid (too few rows/columns)")

    # Column position IS the source of truth for local_idx (exactly as
    # genMaster.py writes it): col 2 -> local_idx 0, col 3 -> local_idx 1, ...
    n_cols = max_col - 1  # excluding the "Day \ Time" column
    if n_cols != SLOTS_PER_DAY:
        raise ValueError(
            f"Uploaded file has {n_cols} timeslot column(s), expected {SLOTS_PER_DAY}. "
            "Don't add, remove, or reorder columns — only edit the text inside class cells."
        )

    for local_idx in range(SLOTS_PER_DAY):
        ci = local_idx + 2
        header_val = ws.cell(row=1, column=ci).value
        header_start = _parse_header_time_label(header_val)
        expected_start = _DAY_TEMPLATE[local_idx]["start"]
        if header_start != expected_start:
            raise ValueError(
                f"Column {ci} header is '{header_val}', expected a "
                f"{expected_start} slot. Don't reorder timeslot columns."
            )

    # Forward-fill the merged "Day" column (col 1) so every row knows which
    # day block it belongs to.
    row_day = {}
    current_day = None
    for r in range(2, max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val:
            candidate = str(val).strip()
            if candidate not in DAYS_ORDER:
                raise ValueError(
                    f"Row {r}, column A has '{val}', which isn't one of {DAYS_ORDER}. "
                    "Don't edit the Day column."
                )
            current_day = candidate
        if current_day is None:
            raise ValueError(f"Row {r} has no Day value above it — the Day column looks corrupted.")
        row_day[r] = current_day

    # Walk every non-lunch cell, skip blanks and "-> ... (contd.)" markers.
    entries = []
    for r in range(2, max_row + 1):
        day = row_day[r]
        for local_idx in range(SLOTS_PER_DAY):
            if local_idx == _LUNCH_LOCAL_IDX:
                continue
            ci = local_idx + 2
            value = ws.cell(row=r, column=ci).value
            if value is None:
                continue
            text = str(value).strip()
            if not text or text.startswith("\u21b3"):
                continue  # blank, or a practical's "contd." marker

            lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
            if len(lines) < 3:
                raise ValueError(
                    f"Cell at row {r}, {day} column {ci} doesn't look like a class "
                    f"entry (expected 3 lines: code/type, title, room/teacher). "
                    f"Got: {text!r}"
                )

            header_match = _HEADER_LINE_RE.match(lines[0])
            if not header_match:
                raise ValueError(
                    f"Cell at row {r}, {day} column {ci}: couldn't parse '{lines[0]}' "
                    "as 'COURSE_CODE [TYPE]'."
                )
            course_code, type_abbrev = header_match.group(1), header_match.group(2).upper()
            unit_type = TYPE_FROM_ABBREV.get(type_abbrev)
            if unit_type is None:
                raise ValueError(
                    f"Cell at row {r}, {day} column {ci}: unknown class type "
                    f"'[{type_abbrev}]'. Must be one of LEC, TUT, PRAC."
                )

            room_match = _ROOM_LINE_RE.match(lines[-1])
            if not room_match:
                raise ValueError(
                    f"Cell at row {r}, {day} column {ci}: couldn't parse '{lines[-1]}' "
                    "as 'Room <room_id> \u00b7 <teacher>'."
                )
            room_id, teacher = room_match.group(1).strip(), room_match.group(2).strip()

            entries.append({
                "day": day,
                "local_idx": local_idx,
                "course_code": course_code,
                "unit_type": unit_type,
                "room_id": room_id,
                "teacher": teacher,
            })

    return entries


def import_master_timetable(
    input_xlsx: str,
    extracted_json: str,
    output_json: str,
) -> dict:
    """Rebuild timetable.json from an admin-edited copy of the merged-cell
    master_timetable.xlsx grid. Classes may be freely added, removed, or
    retyped — see module docstring for how enrollment is kept correct
    despite that.
    """
    if not os.path.exists(input_xlsx):
        raise FileNotFoundError(f"Uploaded file not found: {input_xlsx}")
    if not os.path.exists(extracted_json):
        raise FileNotFoundError(
            f"Cannot import: '{extracted_json}' does not exist. Generate a "
            "timetable at least once (so room/student/curriculum data exists) "
            "before uploading a hand-edited master timetable."
        )

    with open(extracted_json, encoding="utf-8") as f:
        extracted = json.load(f)

    units = extracted["units"]
    subjects = extracted["subjects"]
    rooms = extracted["rooms"]
    students = extracted["students"]
    groups = extracted["enrollment_groups"]

    # enrollment_groups is stored as unit_id lists (how extractor.py
    # originally built it). Collapse each group down to the set of
    # course_codes it covers — this stays valid even after the sheet
    # adds/removes/retypes sessions, since a course's presence in a
    # group's curriculum doesn't change just because its session count did.
    group_courses = {}
    for key, uids in groups.items():
        group_courses[key] = {units[uid]["course_code"] for uid in uids if uid in units}

    entries = _read_master_grid(input_xlsx)
    if not entries:
        raise ValueError("Uploaded file has no class cells to import")

    day_rank = {d: i for i, d in enumerate(DAYS_ORDER)}
    entries.sort(key=lambda e: (e["course_code"], e["unit_type"], day_rank[e["day"]], e["local_idx"]))

    # Assign a fresh unit_id per parsed cell: CODE_L1, CODE_L2, CODE_P1, ...
    # Generated fresh on every import, so nothing depends on an id staying
    # the same across uploads.
    seen_count = defaultdict(int)

    classes = []
    for entry in entries:
        code = entry["course_code"]
        unit_type = entry["unit_type"]
        seen_count[(code, unit_type)] += 1
        uid = f"{code}_{TYPE_LETTER[unit_type]}{seen_count[(code, unit_type)]}"

        room = rooms.get(entry["room_id"])
        if not room:
            raise ValueError(
                f"Class cell for {code} [{TYPE_ABBREV.get(unit_type, unit_type)}] "
                f"on {entry['day']} references unknown room_id '{entry['room_id']}'"
            )

        local_idx = entry["local_idx"]
        slot_index = day_rank[entry["day"]] * SLOTS_PER_DAY + local_idx
        slot1 = ALL_SLOTS[slot_index]
        slots = 2 if unit_type == "practical" else 1

        if slots == 2:
            if local_idx + 1 >= SLOTS_PER_DAY:
                raise ValueError(
                    f"Class cell for {code} [PRAC] on {entry['day']} starts in the "
                    "last timeslot of the day, leaving no room for its 2nd half. "
                    "Move it earlier in the day."
                )
            end_time = ALL_SLOTS[slot_index + 1]["end"]
        else:
            end_time = slot1["end"]

        subj = subjects.get(code, {})
        classes.append({
            "unit_id":     uid,
            "course_code": code,
            "title":       subj.get("title", code),
            "credits":     subj.get("credits", 0),
            "type":        unit_type,
            "slots":       slots,
            "day":         entry["day"],
            "start":       slot1["start"],
            "end":         end_time,
            "slot_index":  slot_index,
            "room_id":     entry["room_id"],
            "room_type":   room.get("room_type", ""),
            "capacity":    room.get("capacity", 0),
            "teacher":     entry["teacher"],
        })

    classes.sort(key=lambda c: (day_rank[c["day"]], c["slot_index"]))
    by_day = {day: [c for c in classes if c["day"] == day] for day in DAYS_ORDER}

    by_student = {}
    for enr, s in students.items():
        key = f"sem{s['semester']}|{s['major']}|{s['minor']}"
        my_courses = group_courses.get(key, set())
        my_cls = [c for c in classes if c["course_code"] in my_courses]
        by_student[enr] = {
            "enrollment_no": enr,
            "name":          s["name"],
            "email":         s["email"],
            "major":         s["major"],
            "minor":         s["minor"],
            "semester":      s["semester"],
            "classes":       my_cls,
        }

    out = {
        "meta": {
            "total_classes":         len(classes),
            "total_students":        len(by_student),
            "days":                  DAYS_ORDER,
            "slots_per_day":         SLOTS_PER_DAY,
            "slot_duration_minutes": _SLOT_MINUTES,
            "timeslots":             ALL_SLOTS,
            "source":                "manual_master_upload",
        },
        "timetable":  classes,
        "by_day":     by_day,
        "by_student": by_student,
    }

    os.makedirs(os.path.dirname(output_json), exist_ok=True)
    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    return out


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    OUTPUT_DIR = os.path.join(BASE_DIR, "output")

    if len(sys.argv) < 2:
        print("Usage:")
        print("  python master_sync.py <edited_master_timetable.xlsx>")
        raise SystemExit(1)

    result = import_master_timetable(
        sys.argv[1],
        os.path.join(OUTPUT_DIR, "extracted_data.json"),
        os.path.join(OUTPUT_DIR, "timetable.json"),
    )
    print(f"  \u2713 timetable.json rebuilt from {sys.argv[1]}")
    print(f"  Classes  : {result['meta']['total_classes']}")
    print(f"  Students : {result['meta']['total_students']}")