"""
extractor.py  v3
─────────────────────────────────────────────────────────────────────────────
Place this in: backend_new/
Excel files go in: backend_new/data/
  - backend_new/data/software_files.xlsx
  - backend_new/data/rooms.xlsx

KEY CHANGE FROM v2:
  Subjects now expand into schedulable UNITS based on L, T, P columns:
    - Each L (lecture)  → 1 unit,  1 slot  (50 min)
    - Each T (tutorial) → 1 unit,  1 slot  (50 min)
    - Each P (practical) → 1 unit, 2 slots (100 min, consecutive)

  A subject TB2401 with L=4, T=0, P=0 produces 4 units:
    TB2401_L1, TB2401_L2, TB2401_L3, TB2401_L4
  Each unit is scheduled independently — can land on different days,
  different rooms, anywhere across the week.

  The conflict graph and scheduler operate on unit_ids, not course_codes.
"""

import openpyxl
import json
import os
from collections import defaultdict
from dataclasses import dataclass, asdict

# ─────────────────────────────────────────────────────────────────────────────
# Canonical program name map
# ─────────────────────────────────────────────────────────────────────────────
PROGRAM_NAME_MAP = {
    "data science & ai":               "Data Science & AI",
    "data science and ai":             "Data Science & AI",
    "data science":                    "Data Science & AI",
    "cs/dsai":                         "Data Science & AI",
    "computer science":                "Computer Science",
    "computer sciences":               "Computer Science",
    "psychology":                      "Psychology",
    "b.a psychology":                  "Psychology",
    "b.a psychology hons.":            "Psychology",
    "psychology/cognitive science":    "Psychology",
    "political science":               "Political Science",
    "political science,":              "Political Science",
    "philosophy/political science":    "Political Science",
    "political science/economics":     "Political Science",
    "environment & sustainability":    "Environment & Sustainability",
    "environment and sustainability":  "Environment & Sustainability",
    "philosophy":                      "Philosophy",
    "casp":                            "CASP",
    "casp ":                           "CASP",
    "casp (art & media)":              "CASP",
    "arts & media":                    "CASP",
    "arts & media   ":                 "CASP",
    "lcs":                             "LCS",
    "literary and cultural studies":   "LCS",
    "sociology":                       "Sociology",
    "sociology ":                      "Sociology",
    "bba":                             "BBA",
    "economics":                       "Economics",
    "history":                         "History",
    "mathematics":                     "Mathematics",
    "physics":                         "Physics",
    "biotechnology":                   "Biotechnology",
    "cognitive science":               "Cognitive Science",
    "b com":                           "B Com",
    "chemistry":                       "Chemistry",
    "chemsitry":                       "Chemistry",
}

MAX_P_PER_WEEK   = 4   # cap extreme/erroneous P values, with a warning
VALID_OFFERED_AS = {"major", "minor", "both", "major/minor"}


def _get_sheet(wb, preferred_name):
    """Return wb[preferred_name] if that sheet exists, else the workbook's
    active/first sheet. Lets us accept uploaded files whose sheet tab name
    doesn't exactly match what the original combined workbook used, as long
    as the COLUMN layout is unchanged."""
    if preferred_name in wb.sheetnames:
        return wb[preferred_name]
    return wb.active


def _canonical(name):
    if not name or str(name).strip() in ("", "nan", "None"):
        return ""
    key = str(name).strip().lower()
    if key not in PROGRAM_NAME_MAP:
        print(f"    ⚠  Unknown program '{name}' — keeping as-is. Add to PROGRAM_NAME_MAP.")
        return str(name).strip()
    return PROGRAM_NAME_MAP[key]


def _norm_offered(val):
    if not val or str(val).strip() in ("", "nan", "None"):
        return ""
    n = str(val).strip().lower()
    if n == "both":
        n = "major/minor"
    return n


# ─────────────────────────────────────────────────────────────────────────────
# Dataclasses
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class Subject:
    course_code: str
    title:       str
    L: int
    T: int
    P: int
    credits: int

@dataclass
class SchedulableUnit:
    unit_id:     str    # e.g. "TB2401_L1", "TA2405_P2"
    course_code: str
    unit_type:   str    # "lecture" | "tutorial" | "practical"
    slot_count:  int    # 1 for lecture/tutorial, 2 for practical

@dataclass
class Teacher:
    teacher_code: str
    teacher_name: str
    courses:      list

@dataclass
class Room:
    room_id:   str
    capacity:  int
    room_type: str

@dataclass
class Student:
    enrollment_no: str
    name:     str
    email:    str
    major:    str
    minor:    str
    semester: int

@dataclass
class CurriculumEntry:
    program:      str
    semester:     int
    course_code:  str
    title:        str
    credits:      int
    offered_as:   str
    teacher:      str
    teacher_code: str


# ─────────────────────────────────────────────────────────────────────────────
# Unit expansion
# ─────────────────────────────────────────────────────────────────────────────

def _expand_to_units(subjects):
    """Expands each Subject into SchedulableUnits based on L, T, P."""
    units = {}
    skipped_zero = []
    skipped_cap  = []

    for code, subj in subjects.items():
        L, T, P = subj.L, subj.T, subj.P

        if P > MAX_P_PER_WEEK:
            skipped_cap.append(f"{code} (P={P} → capped to {MAX_P_PER_WEEK})")
            P = MAX_P_PER_WEEK

        if L + T + P == 0:
            skipped_zero.append(code)
            continue

        for i in range(1, L + 1):
            uid = f"{code}_L{i}"
            units[uid] = SchedulableUnit(uid, code, "lecture", 1)
        for i in range(1, T + 1):
            uid = f"{code}_T{i}"
            units[uid] = SchedulableUnit(uid, code, "tutorial", 1)
        for i in range(1, P + 1):
            uid = f"{code}_P{i}"
            units[uid] = SchedulableUnit(uid, code, "practical", 2)

    if skipped_zero:
        print(f"    ⚠  {len(skipped_zero)} subjects with L=T=P=0 (e.g. dissertations) — no timetable entry: {skipped_zero}")
    if skipped_cap:
        print(f"    ⚠  Capped extreme P values: {skipped_cap}")

    return units


# ─────────────────────────────────────────────────────────────────────────────
# Loaders
# ─────────────────────────────────────────────────────────────────────────────

def _load_curriculum_and_teachers(curriculum_wb, teacher_name_wb, teachers_wb):
    print("\n  [1/3] Loading curriculum + teachers ...")

    ws        = _get_sheet(curriculum_wb, 'curriculum')
    rows      = list(ws.iter_rows(values_only=True))
    data_rows = [r for r in rows[2:] if any(c is not None for c in r)]

    curriculum       = defaultdict(list)
    subjects         = {}
    teacher_subjects_raw = defaultdict(list)   # course_code -> [teacher names], ordered & deduped
    teacher_courses  = defaultdict(set)
    teacher_info     = {}

    def _add_teacher(code, name):
        if name and name not in teacher_subjects_raw[code]:
            teacher_subjects_raw[code].append(name)

    for r in data_rows:
        program = _canonical(str(r[1]).strip() if r[1] else "")

        raw_sem = r[2]
        if hasattr(raw_sem, 'day'):
            semester = raw_sem.day
        else:
            try:
                semester = int(float(str(raw_sem)))
            except Exception:
                print(f"    ⚠  Cannot parse semester '{raw_sem}' — skipping row")
                continue

        code         = str(r[3]).strip() if r[3] else ""
        title        = str(r[4]).strip() if r[4] else ""
        L            = int(r[5] or 0)
        T            = int(r[6] or 0)
        P            = int(r[7] or 0)
        credits      = int(r[8] or 0)
        offered_raw  = str(r[9]).strip() if r[9] else ""
        teacher_name = str(r[10]).strip() if r[10] else ""
        teacher_code = str(r[11]).strip() if r[11] else ""

        if not program or not code:
            continue

        offered = _norm_offered(offered_raw)
        if offered not in VALID_OFFERED_AS:
            print(f"    ⚠  Invalid offered_as='{offered_raw}' for {code} — skipping")
            continue

        entry = CurriculumEntry(program, semester, code, title, credits,
                                 offered, teacher_name, teacher_code)
        curriculum[(program, semester)].append(entry)

        if code not in subjects:
            subjects[code] = Subject(code, title, L, T, P, credits)

        if teacher_name and teacher_code:
            _add_teacher(code, teacher_name)
            teacher_info[teacher_code] = teacher_name
            teacher_courses[teacher_code].add(code)

    ws_tn = _get_sheet(teacher_name_wb, 'teacher_ name')
    for r in list(ws_tn.iter_rows(values_only=True))[1:]:
        if not any(c for c in r): continue
        code = str(r[0]).strip() if r[0] else ""
        name = str(r[1]).strip() if r[1] else ""
        tc   = str(r[2]).strip() if r[2] else ""
        if not code or not name: continue
        _add_teacher(code, name)
        if tc:
            teacher_info[tc] = name
            teacher_courses[tc].add(code)

    ws_t = _get_sheet(teachers_wb, 'teachers')
    for r in list(ws_t.iter_rows(values_only=True))[1:]:
        if not any(c for c in r): continue
        code = str(r[0]).strip() if r[0] else ""
        name = str(r[1]).strip() if r[1] else ""
        tc   = str(r[2]).strip() if r[2] else ""
        if not code or not name or code == "nan" or name == "nan": continue
        # This sheet is a fallback source — only use it for codes that got
        # no teacher at all from the more authoritative sheets above.
        if code not in teacher_subjects_raw:
            _add_teacher(code, name)
        if tc and tc not in ("nan", ""):
            teacher_info[tc] = name
            teacher_courses[tc].add(code)

    teachers = {
        tc: Teacher(tc, name, sorted(teacher_courses[tc]))
        for tc, name in teacher_info.items()
    }

    # Collapse each code's teacher list into a single display/scheduling
    # string, e.g. "Dr. A Sharma & Dr. B Mehta". Downstream code (the SA
    # scheduler's teacher-clash check, and the JSON "teacher" field shown
    # in index.html) treats this as one atomic value per subject — which
    # also means all teachers of a shared-teaching subject correctly get
    # treated as unavailable at the same time in the scheduler.
    teacher_subjects = {
        code: " & ".join(names) for code, names in teacher_subjects_raw.items()
    }

    curriculum = dict(curriculum)
    print(f"    ✓ {len(subjects)} unique subjects | "
          f"{sum(len(v) for v in curriculum.values())} curriculum entries | "
          f"{len(curriculum)} (program, semester) combinations")
    print(f"    ✓ {len(teachers)} teachers | {len(teacher_subjects)} subject→teacher mappings")
    multi_teacher = {c: n for c, n in teacher_subjects_raw.items() if len(n) > 1}
    if multi_teacher:
        print(f"    ℹ  {len(multi_teacher)} subjects have multiple teachers:")
        for c, names in sorted(multi_teacher.items()):
            print(f"      - {c}: {' & '.join(names)}")
    return curriculum, subjects, teachers, teacher_subjects


def _load_rooms(rooms_path):
    print("\n  [2/3] Loading rooms.xlsx ...")
    wb = openpyxl.load_workbook(rooms_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))[1:]

    rooms = {}
    for r in rows:
        if not any(c for c in r): continue
        rid = str(r[0]).strip() if r[0] else ""
        cap = 0
        if len(r) > 3 and r[3] is not None:
            try: cap = int(float(str(r[3])))
            except Exception: cap = 0
        if not rid or rid.lower() in ("nan", "room id", "room_id", ""): continue
        prefix    = rid.replace("-", "").replace(" ", "").upper()[:2]
        room_type = "lab" if prefix == "CL" else "lecture"
        if rid not in rooms:
            rooms[rid] = Room(rid, cap, room_type)

    lh = sum(1 for r in rooms.values() if r.room_type == "lecture")
    cl = sum(1 for r in rooms.values() if r.room_type == "lab")
    print(f"    ✓ {len(rooms)} rooms  ({lh} lecture halls, {cl} computer labs)")
    return rooms


def _load_students(students_wb):
    print("\n  [3/3] Loading students ...")
    ws   = _get_sheet(students_wb, 'Sheet1')
    rows = list(ws.iter_rows(values_only=True))[1:]

    students   = {}
    skipped    = 0
    sem_counts = defaultdict(int)

    for r in rows:
        if not any(c for c in r): continue
        enr = r[0]
        if enr is None:
            skipped += 1; continue
        try:
            enr = str(int(float(str(enr))))
        except Exception:
            skipped += 1; continue

        name  = str(r[1]).strip() if r[1] else ""
        email = str(r[2]).strip() if r[2] else ""
        major = _canonical(str(r[3]).strip() if r[3] else "")
        minor = _canonical(str(r[4]).strip() if r[4] else "")
        try: semester = int(float(str(r[5])))
        except Exception: semester = 0

        if enr in students:
            print(f"    ⚠  Duplicate enrollment_no {enr} — skipping")
            continue

        students[enr] = Student(enr, name, email, major, minor, semester)
        sem_counts[semester] += 1

    if skipped:
        print(f"    ⚠  Skipped {skipped} blank/invalid rows")
    print(f"    ✓ {len(students)} students loaded")
    for sem, cnt in sorted(sem_counts.items()):
        print(f"      Semester {sem}: {cnt} students")
    return students


# ─────────────────────────────────────────────────────────────────────────────
# Enrollment groups (unit-based)
# ─────────────────────────────────────────────────────────────────────────────

def _build_enrollment_groups(students, curriculum, units):
    print("\n  Building enrollment groups ...")

    code_to_units = defaultdict(list)
    for uid, unit in units.items():
        code_to_units[unit.course_code].append(uid)

    combos = set()
    for s in students.values():
        combos.add((s.semester, s.major, s.minor))

    groups  = {}
    missing = set()

    for (semester, major, minor) in sorted(combos):
        course_codes = []

        major_entries = curriculum.get((major, semester), [])
        if not major_entries and major:
            missing.add(f"major='{major}' sem={semester}")
        for e in major_entries:
            if e.offered_as in ("major", "major/minor"):
                course_codes.append(e.course_code)

        if minor:
            minor_entries = curriculum.get((minor, semester), [])
            if not minor_entries:
                missing.add(f"minor='{minor}' sem={semester}")
            for e in minor_entries:
                if e.offered_as in ("minor", "major/minor"):
                    if e.course_code not in course_codes:
                        course_codes.append(e.course_code)

        course_codes = list(dict.fromkeys(course_codes))

        unit_ids = []
        for code in course_codes:
            unit_ids.extend(code_to_units.get(code, []))

        groups[(semester, major, minor)] = unit_ids

    if missing:
        print(f"    ⚠  No curriculum data for:")
        for m in sorted(missing): print(f"      - {m}")

    active = sum(1 for v in groups.values() if v)
    print(f"    ✓ {len(groups)} enrollment groups total  |  {active} active (>0 units)")
    for (sem, maj, mn), uids in sorted(groups.items()):
        if uids:
            label = f"sem{sem} | {maj}" + (f" + {mn}" if mn else "")
            print(f"      {label:55s} → {len(uids)} slot-entries")

    return groups


def _build_conflict_graph(groups):
    print("\n  Building conflict graph ...")
    graph = defaultdict(set)
    for unit_ids in groups.values():
        n = len(unit_ids)
        for i in range(n):
            for j in range(i + 1, n):
                a, b = unit_ids[i], unit_ids[j]
                graph[a].add(b)
                graph[b].add(a)
    graph = dict(graph)
    edges = sum(len(v) for v in graph.values()) // 2
    print(f"    ✓ {len(graph)} units in conflict graph  |  {edges} conflict edges")
    return graph


def _validate(students, curriculum, subjects, teacher_subjects, units):
    print("\n  Running validation ...")
    all_programs = {prog for (prog, _) in curriculum.keys()}

    no_teacher = [f"{c} ({subjects[c].title})"
                  for c in subjects if c not in teacher_subjects
                  and (subjects[c].L + subjects[c].T + subjects[c].P) > 0]
    if no_teacher:
        print(f"    ⚠  {len(no_teacher)} schedulable subjects with no teacher:")
        for s in no_teacher: print(f"      - {s}")

    missing_prog = set()
    for s in students.values():
        if s.major and s.major not in all_programs:
            missing_prog.add(f"major '{s.major}' sem {s.semester}")
        if s.minor and s.minor not in all_programs:
            missing_prog.add(f"minor '{s.minor}' sem {s.semester}")
    if missing_prog:
        print(f"    ⚠  Programs in students not found in curriculum ({len(missing_prog)}):")
        for p in sorted(missing_prog): print(f"      - {p}")

    lec  = sum(1 for u in units.values() if u.unit_type == "lecture")
    tut  = sum(1 for u in units.values() if u.unit_type == "tutorial")
    prac = sum(1 for u in units.values() if u.unit_type == "practical")
    print(f"    ✓ {len(units)} total schedulable units  "
          f"({lec} lectures + {tut} tutorials + {prac} practicals)")
    print(f"    ✓ Total slot-entries needed: {lec + tut + prac*2}  "
          f"(available: 55 slots/week × N rooms)")

    if not no_teacher and not missing_prog:
        print("    ✓ All checks passed.")


def _to_json(data):
    def conv(obj):
        if hasattr(obj, "__dataclass_fields__"): return asdict(obj)
        if isinstance(obj, set):                 return sorted(list(obj))
        if isinstance(obj, dict):                return {str(k): conv(v) for k, v in obj.items()}
        if isinstance(obj, list):                return [conv(i) for i in obj]
        return obj

    return {
        "subjects":    conv(data["subjects"]),
        "units":       conv(data["units"]),
        "teachers":    conv(data["teachers"]),
        "rooms":       conv(data["rooms"]),
        "students":    conv(data["students"]),
        "curriculum": {
            f"{prog}|sem{sem}": [asdict(e) for e in entries]
            for (prog, sem), entries in data["curriculum"].items()
        },
        "teacher_subjects": data["teacher_subjects"],
        "enrollment_groups": {
            f"sem{sem}|{maj}|{mn}": uids
            for (sem, maj, mn), uids in data["enrollment_groups"].items()
        },
        "conflict_graph": {
            k: sorted(v) for k, v in data["conflict_graph"].items()
        },
    }


def load_data(students_xlsx, curriculum_xlsx, teacher_name_xlsx, teachers_xlsx,
              rooms_xlsx, output_json, active_semesters=None):
    print("=" * 65)
    print("  TSLAS DATA EXTRACTOR  v3")
    print("=" * 65)

    curriculum_wb   = openpyxl.load_workbook(curriculum_xlsx, read_only=True, data_only=True)
    teacher_name_wb = openpyxl.load_workbook(teacher_name_xlsx, read_only=True, data_only=True)
    teachers_wb     = openpyxl.load_workbook(teachers_xlsx, read_only=True, data_only=True)
    students_wb     = openpyxl.load_workbook(students_xlsx, read_only=True, data_only=True)

    curriculum, subjects, teachers, teacher_subjects = \
        _load_curriculum_and_teachers(curriculum_wb, teacher_name_wb, teachers_wb)
    rooms    = _load_rooms(rooms_xlsx)
    students = _load_students(students_wb)

    if active_semesters:
        before   = len(students)
        students = {enr: s for enr, s in students.items()
                    if s.semester in active_semesters}
        filtered = before - len(students)
        parity   = 'even' if active_semesters[0] % 2 == 0 else 'odd'
        print(f"  Semester filter: {parity.upper()} {active_semesters}")
        print(f"  Students after filter: {len(students)}  ({filtered} excluded)")

    print("\n  Expanding subjects into schedulable units ...")
    units = _expand_to_units(subjects)

    _validate(students, curriculum, subjects, teacher_subjects, units)

    enrollment_groups = _build_enrollment_groups(students, curriculum, units)
    conflict_graph    = _build_conflict_graph(enrollment_groups)

    data = {
        "subjects":          subjects,
        "units":             units,
        "teachers":          teachers,
        "rooms":             rooms,
        "students":          students,
        "curriculum":        curriculum,
        "teacher_subjects":  teacher_subjects,
        "enrollment_groups": enrollment_groups,
        "conflict_graph":    conflict_graph,
    }

    os.makedirs(os.path.dirname(output_json), exist_ok=True)
    with open(output_json, "w", encoding="utf-8") as f:
        json.dump(_to_json(data), f, indent=2, ensure_ascii=False)
    print(f"\n  ✓ Saved → {output_json}")

    print("\n" + "=" * 65)
    print("  EXTRACTION COMPLETE")
    print("=" * 65)
    parity_label = ""
    if active_semesters:
        p = 'EVEN' if active_semesters[0] % 2 == 0 else 'ODD'
        parity_label = f"  ({p}: {active_semesters})"
    lec  = sum(1 for u in units.values() if u.unit_type == "lecture")
    tut  = sum(1 for u in units.values() if u.unit_type == "tutorial")
    prac = sum(1 for u in units.values() if u.unit_type == "practical")
    print(f"  Subjects          : {len(subjects)}")
    print(f"  Schedulable units : {len(units)}  "
          f"({lec} lectures + {tut} tutorials + {prac} practicals)")
    print(f"  Slot-entries/week : {lec + tut + prac*2}")
    print(f"  Teachers          : {len(teachers)}")
    print(f"  Rooms             : {len(rooms)}")
    print(f"  Students          : {len(students)}{parity_label}")
    active = sum(1 for v in enrollment_groups.values() if v)
    print(f"  Enrollment groups : {len(enrollment_groups)}  ({active} active)")
    edges = sum(len(v) for v in conflict_graph.values()) // 2
    print(f"  Conflict edges    : {edges}")
    print("=" * 65)

    return data