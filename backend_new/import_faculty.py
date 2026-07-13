"""Import faculty records from the master Excel workbook into DynamoDB.

Safe to re-run: profile fields (name/designation/email/teacher_code) are
always refreshed, but a faculty member's password is only set the first
time their employee code is seen, so re-running this after someone has
changed their password will NOT reset it.

Usage:
    python import_faculty.py --file "../Frontend/data/FACULTY DETAILS - TSLAS.xlsx"
    python import_faculty.py --file path/to/file.xlsx --dry-run
"""

from __future__ import annotations

import argparse
import os
import sys
from typing import Iterable

import bcrypt
import openpyxl
from botocore.exceptions import ClientError
from dotenv import load_dotenv

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

load_dotenv(os.path.join(BASE_DIR, ".env"))

from db import get_faculty_table  # noqa: E402
from excel_utils import cell, find_column  # noqa: E402

EMPLOYEE_CODE_HEADERS = {"empolyee code", "employee code", "emp code", "employee no"}
TEACHER_CODE_HEADERS = {"teacher code"}
NAME_HEADERS = {"name of faculty name", "name of faculty", "faculty name", "name"}
DESIGNATION_HEADERS = {"designation"}
EMAIL_HEADERS = {"email", "email address"}


def iter_faculty_rows(path: str) -> Iterable[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration:
            continue

        employee_code_idx = find_column(headers, EMPLOYEE_CODE_HEADERS)
        name_idx = find_column(headers, NAME_HEADERS)
        if employee_code_idx is None or name_idx is None:
            continue

        teacher_code_idx = find_column(headers, TEACHER_CODE_HEADERS)
        designation_idx = find_column(headers, DESIGNATION_HEADERS)
        email_idx = find_column(headers, EMAIL_HEADERS)

        for row in rows:
            if not row:
                continue
            employee_code = cell(row, employee_code_idx)
            name = cell(row, name_idx)
            if not employee_code or not name:
                continue

            yield {
                "employee_code": employee_code,
                "name": name,
                "teacher_code": cell(row, teacher_code_idx),
                "designation": cell(row, designation_idx),
                "email": cell(row, email_idx),
                "source_sheet": sheet_name,
            }


def import_faculty(path: str, default_password: str, dry_run: bool = False) -> None:
    deduped: dict[str, dict[str, str]] = {}
    for record in iter_faculty_rows(path):
        deduped[record["employee_code"]] = record  # later sheets win on duplicate employee codes

    print(f"Parsed {len(deduped)} unique faculty members from '{path}'")

    if dry_run:
        for record in list(deduped.values())[:5]:
            print(record)
        print("Dry run: nothing was written to DynamoDB.")
        return

    default_password_hash = bcrypt.hashpw(default_password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    table = get_faculty_table()

    written = 0
    for record in deduped.values():
        try:
            table.update_item(
                Key={"employee_code": record["employee_code"]},
                UpdateExpression=(
                    "SET #name = :name, teacher_code = :teacher_code, designation = :designation, "
                    "email = :email, source_sheet = :sheet, "
                    "password_hash = if_not_exists(password_hash, :pwhash), "
                    "password_is_default = if_not_exists(password_is_default, :pwdefault)"
                ),
                ExpressionAttributeNames={"#name": "name"},
                ExpressionAttributeValues={
                    ":name": record["name"],
                    ":teacher_code": record["teacher_code"],
                    ":designation": record["designation"],
                    ":email": record["email"],
                    ":sheet": record["source_sheet"],
                    ":pwhash": default_password_hash,
                    ":pwdefault": True,
                },
            )
        except ClientError as exc:
            print(f"Failed to write {record['employee_code']}: {exc}")
            continue

        written += 1

    print(f"Done. Wrote/updated {written} faculty members in DynamoDB table '{table.table_name}'.")
    print(f"New faculty default to password '{default_password}' and must change it on first login.")


def main() -> None:
    parser = argparse.ArgumentParser(description="Import faculty from Excel into DynamoDB")
    parser.add_argument("--file", required=True, help="Path to the faculty Excel workbook")
    parser.add_argument("--default-password", default="tiet12345", help="Password assigned to newly-imported faculty")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print a sample without writing to DynamoDB")
    args = parser.parse_args()

    import_faculty(args.file, args.default_password, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
