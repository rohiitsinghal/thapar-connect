"""Import student records from the master Excel workbook into DynamoDB.

Safe to re-run: profile fields (name/email/major/minor) are always refreshed,
but a student's password is only set the first time their roll number is
seen, so re-running this after students have changed their password will
NOT reset it.

Usage:
    python import_students.py --file "../Frontend/data/Students List  (4).xlsx"
    python import_students.py --file path/to/file.xlsx --dry-run
"""

from __future__ import annotations

import argparse
import os
import sys
from typing import Any, Iterable, Optional

import bcrypt
import openpyxl
from botocore.exceptions import ClientError
from dotenv import load_dotenv

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

load_dotenv(os.path.join(BASE_DIR, ".env"))

from db import get_students_table  # noqa: E402

ROLL_HEADERS = {"roll no", "roll no.", "roll number", "enrollment number", "enrollment no", "enrollment no.", "enrollmentno"}
NAME_HEADERS = {"name", "student name"}
EMAIL_HEADERS = {"email address", "email", "email id", "students email address"}
MAJOR_HEADERS = {"major", "major "}
MINOR_HEADERS = {"minor"}


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _find_column(headers: list[Any], candidates: set[str]) -> Optional[int]:
    for index, header in enumerate(headers):
        if _normalize_header(header) in candidates:
            return index
    return None


def _cell(row: tuple[Any, ...], index: Optional[int]) -> str:
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def iter_student_rows(path: str) -> Iterable[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration:
            continue

        roll_idx = _find_column(headers, ROLL_HEADERS)
        name_idx = _find_column(headers, NAME_HEADERS)
        if roll_idx is None or name_idx is None:
            continue

        email_idx = _find_column(headers, EMAIL_HEADERS)
        major_idx = _find_column(headers, MAJOR_HEADERS)
        minor_idx = _find_column(headers, MINOR_HEADERS)

        for row in rows:
            if not row:
                continue
            roll_no = _cell(row, roll_idx).upper()
            name = _cell(row, name_idx)
            if not roll_no or not name:
                continue

            yield {
                "roll_no": roll_no,
                "name": name,
                "email": _cell(row, email_idx),
                "major": _cell(row, major_idx),
                "minor": _cell(row, minor_idx),
                "source_sheet": sheet_name,
            }


def import_students(path: str, default_password: str, dry_run: bool = False) -> int:
    deduped: dict[str, dict[str, str]] = {}
    for record in iter_student_rows(path):
        deduped[record["roll_no"]] = record  # later sheets win on duplicate roll numbers

    print(f"Parsed {len(deduped)} unique students from '{path}'")

    if dry_run:
        for record in list(deduped.values())[:5]:
            print(record)
        print("Dry run: nothing was written to DynamoDB.")
        return 0

    default_password_hash = bcrypt.hashpw(default_password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    table = get_students_table()

    written = 0
    for record in deduped.values():
        try:
            table.update_item(
                Key={"roll_no": record["roll_no"]},
                UpdateExpression=(
                    "SET #name = :name, email = :email, major = :major, minor = :minor, "
                    "source_sheet = :sheet, "
                    "password_hash = if_not_exists(password_hash, :pwhash), "
                    "password_is_default = if_not_exists(password_is_default, :pwdefault)"
                ),
                ExpressionAttributeNames={"#name": "name"},
                ExpressionAttributeValues={
                    ":name": record["name"],
                    ":email": record["email"],
                    ":major": record["major"],
                    ":minor": record["minor"],
                    ":sheet": record["source_sheet"],
                    ":pwhash": default_password_hash,
                    ":pwdefault": True,
                },
            )
        except ClientError as exc:
            print(f"Failed to write {record['roll_no']}: {exc}")
            continue

        written += 1
        if written % 200 == 0:
            print(f"  ...{written} written")

    print(f"Done. Wrote/updated {written} students in DynamoDB table '{table.table_name}'.")
    print(f"New students default to password '{default_password}' and must change it on first login.")
    return written


def main() -> None:
    parser = argparse.ArgumentParser(description="Import students from Excel into DynamoDB")
    parser.add_argument("--file", required=True, help="Path to the students Excel workbook")
    parser.add_argument("--default-password", default="12345", help="Password assigned to newly-imported students")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print a sample without writing to DynamoDB")
    args = parser.parse_args()

    import_students(args.file, args.default_password, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
