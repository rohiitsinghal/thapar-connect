"""Server-side port of Frontend/src/lib/courseData.ts.

Parses the same course-wise workbook the frontend reads, and re-implements
the same fuzzy faculty/student matching rules so that course-material
endpoints can enforce "does this faculty teach this course" / "is this
student enrolled in this course" server-side instead of trusting the client.
"""

from __future__ import annotations

import os
import re
from functools import lru_cache
from typing import Any, Optional

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
COURSE_WORKBOOK_PATH = os.path.join(BASE_DIR, "..", "Frontend", "data", "COURSE WISE SHEET.xlsx")

STOP_WORDS = {"AND", "OF", "THE", "FOR", "IN", "TO", "A", "AN"}

ROMAN_VALUES = {"I": 1, "V": 5, "X": 10}


def _normalize_lookup(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().upper())


def _normalize_phrase(value: str) -> str:
    cleaned = re.sub(r"&", " ", value)
    cleaned = re.sub(r"[^a-z0-9]+", " ", cleaned, flags=re.IGNORECASE)
    tokens = [segment for segment in cleaned.upper().split(" ") if len(segment) > 1 and segment not in STOP_WORDS]
    return " ".join(tokens)


def _normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def _parse_roman_numeral(value: str) -> Optional[int]:
    roman = _normalize_lookup(value)
    total = 0
    previous = 0
    for char in reversed(roman):
        current = ROMAN_VALUES.get(char)
        if not current:
            return None
        if current < previous:
            total -= current
        else:
            total += current
            previous = current
    return total if total > 0 else None


def infer_semester_from_sheet(sheet_name: str) -> Optional[int]:
    normalized_name = sheet_name.strip()

    if re.search(r"foundation", normalized_name, re.IGNORECASE):
        return 1

    explicit_match = re.search(r"\b(\d{1,2})\s*(?:st|nd|rd|th)?\s*sem", normalized_name, re.IGNORECASE)
    if explicit_match:
        return int(explicit_match.group(1))

    roman_match = re.search(r"\b([IVXivx]{1,4})\b", normalized_name)
    if roman_match:
        return _parse_roman_numeral(roman_match.group(1))

    return None


def _escape_regexp(value: str) -> str:
    return re.escape(value)


def _contains_whole_word_phrase(haystack: str, needle: str) -> bool:
    if not needle:
        return False
    return re.search(rf"(?:^|\s){_escape_regexp(needle)}(?:\s|$)", haystack) is not None


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _build_course_search_text(course: dict[str, Any]) -> str:
    parts = [
        course["courseCode"],
        course["title"],
        course["facultyName"],
        course["facultyCode"],
        *course["programs"],
        *course["categories"],
    ]
    return _normalize_phrase(" ".join(part for part in parts if part))


@lru_cache(maxsize=1)
def get_course_catalog() -> list[dict[str, Any]]:
    """Parse the course-wise workbook once per process, mirroring parseCatalog() in courseData.ts."""
    workbook = openpyxl.load_workbook(COURSE_WORKBOOK_PATH, read_only=True, data_only=True)
    course_map: dict[str, dict[str, Any]] = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))

        header_row_index = None
        for index, row in enumerate(rows):
            if any(_normalize_key(_to_text(cell)) == "course code" for cell in row):
                header_row_index = index
                break

        if header_row_index is None:
            continue

        headers = [_to_text(cell) for cell in rows[header_row_index]]
        current_program = ""

        for row in rows[header_row_index + 1 :]:
            if not any(_to_text(cell) for cell in row):
                continue

            record = {headers[i]: _to_text(row[i]) for i in range(min(len(headers), len(row)))}
            course_code = record.get("Course Code", "")
            title = record.get("TITLE") or record.get("Title", "")
            semester_text = record.get("Sem", "")

            if not course_code or not title or not semester_text.strip().isdigit():
                continue
            semester = int(semester_text)

            program = record.get("Program", "") or current_program
            if program:
                current_program = program

            key = f"{_normalize_lookup(course_code)}|{semester}|{_normalize_lookup(title)}"
            credits_text = record.get("Cr", "")
            credits = int(credits_text) if credits_text.strip().isdigit() else 0
            faculty_name = record.get("Faculty Names", "")
            faculty_code = record.get("Faculty Codes", "")
            category = record.get("Major /Minor", "")

            existing = course_map.get(key)
            if existing:
                if program and program not in existing["programs"]:
                    existing["programs"].append(program)
                if category and category not in existing["categories"]:
                    existing["categories"].append(category)
                if not existing["facultyName"] and faculty_name:
                    existing["facultyName"] = faculty_name
                if not existing["facultyCode"] and faculty_code:
                    existing["facultyCode"] = faculty_code
                if not existing["credits"] and credits:
                    existing["credits"] = credits
                continue

            course_map[key] = {
                "courseCode": course_code,
                "title": title,
                "semester": semester,
                "credits": credits,
                "facultyName": faculty_name,
                "facultyCode": faculty_code,
                "programs": [program] if program else [],
                "categories": [category] if category else [],
            }

    return sorted(course_map.values(), key=lambda course: (course["semester"], course["courseCode"]))


def _courses_for_code(course_code: str) -> list[dict[str, Any]]:
    normalized = _normalize_lookup(course_code)
    return [course for course in get_course_catalog() if _normalize_lookup(course["courseCode"]) == normalized]


def _get_faculty_search_terms(faculty_item: dict[str, Any]) -> list[str]:
    raw_terms = [
        faculty_item.get("name", ""),
        faculty_item.get("employee_code", ""),
        faculty_item.get("teacher_code", ""),
        faculty_item.get("email", ""),
    ]
    terms = [_normalize_phrase(term) for term in raw_terms if term]
    return list(dict.fromkeys(term for term in terms if term))


def faculty_teaches_course(course_code: str, faculty_item: dict[str, Any]) -> bool:
    candidate_courses = _courses_for_code(course_code)
    if not candidate_courses:
        return False

    faculty_terms = _get_faculty_search_terms(faculty_item)
    if not faculty_terms:
        return False

    for course in candidate_courses:
        search_text = _build_course_search_text(course)
        if any(_contains_whole_word_phrase(search_text, term) for term in faculty_terms):
            return True

    return False


def _get_student_study_areas(student_item: dict[str, Any]) -> list[str]:
    raw_values: list[str] = []
    for field in ("major", "minor"):
        value = student_item.get(field, "")
        if value:
            raw_values.extend(re.split(r"[/,&|]+", value))

    areas = [_normalize_phrase(value) for value in raw_values]
    return list(dict.fromkeys(area for area in areas if area))


def student_can_view_course(course_code: str, student_item: dict[str, Any]) -> bool:
    candidate_courses = _courses_for_code(course_code)
    if not candidate_courses:
        return False

    semester = infer_semester_from_sheet(student_item.get("source_sheet", ""))
    if not semester:
        # Semester couldn't be inferred: fall back to allowing (mirrors getCoursesForStudent's
        # "no semester -> return full catalog" behaviour) since course_code already narrows scope.
        return True

    semester_matches = [course for course in candidate_courses if course["semester"] == semester]
    if not semester_matches:
        return False

    study_areas = _get_student_study_areas(student_item)
    if not study_areas:
        return True

    for course in semester_matches:
        search_text = _build_course_search_text(course)
        if any(_contains_whole_word_phrase(search_text, area) for area in study_areas):
            return True

    # Mirrors getCoursesForStudent: no study-area match -> fall back to semester match alone.
    return True
